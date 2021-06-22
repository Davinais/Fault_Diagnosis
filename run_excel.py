import openpyxl
import os
from openpyxl.styles import Font, Fill#Connect styles for text
from openpyxl.styles import colors#Con
Path = './Diag_Report/'
workbook = openpyxl.load_workbook('Info_failLog.xlsx')
worksheet = workbook.get_sheet_by_name('工作表1')

for i in range(2,182):
    worksheet.cell(row=i,column=8).value = None
    worksheet.cell(row=i,column=9).value = None
    worksheet.cell(row=i,column=10).value = None
    worksheet.cell(row=i,column=11).value = None


def fine_row (circuit, index):   # calculate fault free k value 
    for i in range(2,182):
        c = worksheet.cell(row=i,column=1).value
        ind = worksheet.cell(row=i,column=2).value
        if ((c == circuit) and (int(ind) == index)):
            return i
    return -1        

def get_candidate ():  
    num = 0
    g = open(Path+file,mode = 'r')
    for line in g: 
        if (line.find("score") != -1):
            num = num + 1
    g.close()
    return num



allFileList = os.listdir(Path)
insertNum = [0]*183

# determine the number of inserted fault in each row
for i in range(2,182):
    if (worksheet.cell(row=i, column=4).value == 'None'):
        insertNum[i] = 1
    elif (worksheet.cell(row=i, column=5).value == 'None'):
        insertNum[i] = 2
    elif (worksheet.cell(row=i, column=6).value == 'None'):
        insertNum[i] = 3
    elif (worksheet.cell(row=i, column=7).value == 'None'):
        insertNum[i] = 4
    else:
        insertNum[i] = 5



for file in allFileList:
    f = open(Path+file,mode = 'r')
    # print(int(file[-14:-12]))
    # print((file[:-16]))
    F = [0]*5
    Row = fine_row(file[:-16],int(file[-14:-12]))  #find the position of this .rpt in excel
    F[0] = worksheet.cell(row=Row,column=3).value
    F[1] = worksheet.cell(row=Row,column=4).value
    F[2] = worksheet.cell(row=Row,column=5).value
    F[3] = worksheet.cell(row=Row,column=6).value
    F[4] = worksheet.cell(row=Row,column=7).value
    if (insertNum[Row] == 1):
        s = F[0].split('/')
        # print(str[0])
        flag = 0
        g = open(Path+file,mode = 'r')
        candidate = len(g.readlines()) - 11  # numbers of candidates
        g.close()
        acc = 0.0
        resolution = 0.0
        detect = 0
        g = open(Path+file,mode = 'r')
      
        for line in g: 
            flag = 0

            
            for l in range(len(s)):
                if (line.find(s[l]) == -1):
                    flag = 1
                    break
            if (line.find("score=100") != -1):
                flag = 0
                break
            if flag==0: 
                break
            # if (line.find("No.5") != -1):
            #     break

        if flag == 0:
            detect = detect + 1
        g.close()

        g = open(Path+file,mode = 'r')
        for line in g: 
            if (line.find("#run time=") != -1):
                worksheet.cell(row=Row, column=11).value = line[10:-2]
                break
        g.close()


        if detect == 1:
            worksheet.cell(row=Row, column=8).value = 'Correct'
            worksheet.cell(row=Row, column=8).font = Font(color= '000000FF')
            acc = 1.0*100
            if (detect!=0): resolution = 1#candidate/detect
            worksheet.cell(row=Row, column=9).value = str(acc) + '%'
            worksheet.cell(row=Row, column=10).value = resolution
    else:
        candidate = 0
        acc = 0.0
        resolution = 0.0
        detect = 0
        _100count = 0
        special_SSF = False
        special_MSF = False
        f.close()
        candidate = get_candidate ()
        


        for i in range(insertNum[Row]):
            s = F[i].split('/')
            flag = 0
            g = open(Path+file,mode = 'r')
            for line in g: 
                flag = 0
                if (line.find("score=100") != -1):
                    _100count = _100count + 1
                    if (_100count > 1) or (candidate == 1):
                        special_SSF = True
                        # detect = candidate
                        break
                if (line.find("Successful MSF!!!") != -1):
                    special_MSF = True
                    # detect = candidate 
                    break

                for l in range(len(s)):
                    if (line.find(s[l]) == -1):
                        flag = 1   # 有其中一項不符合
                        break
                if flag==0: 
                    break
                # if (line.find("No.5") != -1):
                #     break
            if (flag == 0) and (not special_SSF) and (not special_MSF):
                detect = detect + 1
            g.close()

        g = open(Path+file,mode = 'r')
        for line in g: 
            if (line.find("#run time=") != -1):
                worksheet.cell(row=Row, column=11).value = line[10:-2]
                break
        g.close()

        if (special_SSF):
            worksheet.cell(row=Row, column=8).value = 'MSF detected by SSF'
            worksheet.cell(row=Row, column=8).font = Font(color= '00800000')
        elif (special_MSF):
            worksheet.cell(row=Row, column=8).value = 'MSF detected by others'
            worksheet.cell(row=Row, column=8).font = Font(color= '0000CCFF')
        elif (detect == insertNum[Row]):
            worksheet.cell(row=Row, column=8).value = 'MSF Correct'
            worksheet.cell(row=Row, column=8).font = Font(color= '00FF0000')

        

        if (special_SSF or special_MSF): acc = 100
        else: acc = (detect/insertNum[Row])*100

        if (special_SSF or special_MSF): resolution = 1
        elif (detect!=0): resolution = candidate/detect
        else: resolution = "infinite"
        worksheet.cell(row=Row, column=9).value = str(acc) + '%'
        worksheet.cell(row=Row, column=10).value = resolution
        # worksheet.cell(row=Row, column=11).value = detect



        # worksheet.cell(row=Row, column=8).value = 'SSF (perfectly explain faillog result)'
        # worksheet.cell(row=Row, column=8).font = Font(color= '00FF0000')

    f.close()

# for i in range(2,182):
#     if worksheet.cell(row=i, column=8).value == None:
#         worksheet.cell(row=i, column=8).value = 'MSF'
#         worksheet.cell(row=i, column=8).font = Font(color= '00FF00FF')


workbook.save(filename = 'Info_failLog.xlsx')



